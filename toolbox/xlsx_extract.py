"""Extract Excel (.xlsx / .xlsm / .xls) sheets, tables, images to markdown.

Usage:
    python toolbox/xlsx_extract.py <src.xlsx|src.xlsm|src.xls> <out_dir>

Outputs:
    <out_dir>/<stem>_extract.md
    <out_dir>/assets/<stem>/sheetNN_<N>.<ext>   (embedded images per sheet)

Behaviour:
    .xlsx / .xlsm -> parsed directly with openpyxl (cross-platform).
    .xls          -> converted to temp .xlsx via Excel COM (Windows + Office),
                     then parsed the same way.

Each worksheet is emitted as one `## Sheet N — <name>` section containing a
markdown table of the used range (formula cells use cached values when present
via data_only=True). Merged cells keep the value in their top-left only; the
other covered cells are emitted as blank to preserve column alignment.

This tool only does raw extraction; no AI summarisation, no OCR, no charts.
"""
from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


def convert_xls_to_xlsx(src: Path) -> Path:
    """Use Excel COM to convert legacy .xls to .xlsx into a temp file."""
    try:
        import win32com.client  # type: ignore
    except ImportError as e:
        raise SystemExit(
            "pywin32 required for .xls conversion. Run:\n"
            "  & '$env:USERPROFILE\\.venvs\\sb-docs\\Scripts\\python.exe' -m pip install pywin32"
        ) from e

    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)
    tmp = Path(tmp_path)
    tmp.unlink(missing_ok=True)
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(str(src.resolve()), ReadOnly=True)
        # xlOpenXMLWorkbook = 51 (.xlsx)
        wb.SaveAs(str(tmp.resolve()), FileFormat=51)
        wb.Close(SaveChanges=False)
    finally:
        excel.Quit()
    return tmp


def merged_lookup(ws):
    """Build map: (row, col) of non-anchor merged cells -> anchor (row, col)."""
    covered: dict[tuple[int, int], tuple[int, int]] = {}
    for rng in ws.merged_cells.ranges:
        min_r, min_c, max_r, max_c = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if (r, c) != (min_r, min_c):
                    covered[(r, c)] = (min_r, min_c)
    return covered


def used_bounds(ws) -> tuple[int, int]:
    """Return (max_row, max_col) trimmed to last non-empty cell."""
    max_r = 0
    max_c = 0
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value not in (None, ""):
                if cell.row > max_r:
                    max_r = cell.row
                if cell.column > max_c:
                    max_c = cell.column
    return max_r, max_c


def fmt_cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    s = str(v)
    s = s.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "<br>")
    s = s.replace("|", "\\|")
    return s


def render_sheet_table(ws, max_r: int, max_c: int) -> list[str]:
    covered = merged_lookup(ws)
    # Header row: column letters so positional context is preserved.
    header = [get_column_letter(c) for c in range(1, max_c + 1)]
    sep = ["---"] * max_c
    rows_md: list[str] = []
    rows_md.append("| _row_ | " + " | ".join(header) + " |")
    rows_md.append("| --- | " + " | ".join(sep) + " |")
    for r in range(1, max_r + 1):
        cells = []
        for c in range(1, max_c + 1):
            if (r, c) in covered:
                cells.append("")
            else:
                cells.append(fmt_cell(ws.cell(row=r, column=c).value))
        # Skip rows that are entirely blank (after trimming bounds, interior
        # blank rows still emit to preserve row numbers).
        rows_md.append(f"| {r} | " + " | ".join(cells) + " |")
    return rows_md


def extract_sheet_images(ws, assets_dir: Path, sheet_idx: int) -> list[str]:
    """Save embedded raster images for this sheet. Returns saved filenames."""
    saved: list[str] = []
    images = getattr(ws, "_images", []) or []
    for i, img in enumerate(images, 1):
        try:
            blob = img._data() if callable(getattr(img, "_data", None)) else None
            if not blob:
                ref = getattr(img, "ref", None)
                if ref is None:
                    continue
                blob = ref if isinstance(ref, (bytes, bytearray)) else None
            if not blob:
                continue
            fmt = (getattr(img, "format", "") or "png").lower()
            name = f"sheet{sheet_idx:02d}_{i}.{fmt}"
            (assets_dir / name).write_bytes(blob)
            saved.append(name)
        except Exception as e:
            print(f"[xlsx_extract] sheet {sheet_idx} image {i} skipped: {e}")
    return saved


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(2)

    src = Path(sys.argv[1])
    out_dir = Path(sys.argv[2])
    if not src.exists():
        sys.exit(f"source not found: {src}")
    out_dir.mkdir(parents=True, exist_ok=True)

    tmp_to_clean = None
    ext = src.suffix.lower()
    if ext == ".xls":
        print("[xlsx_extract] .xls detected; converting via Excel COM...")
        xlsx_path = convert_xls_to_xlsx(src)
        tmp_to_clean = xlsx_path
    elif ext in (".xlsx", ".xlsm"):
        xlsx_path = src
    else:
        sys.exit(f"unsupported extension: {ext} (expect .xls, .xlsx, .xlsm)")

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    assets_dir = out_dir / "assets" / src.stem
    assets_dir.mkdir(parents=True, exist_ok=True)

    # Sheet overview
    overview_rows = []
    sheet_infos = []
    for idx, name in enumerate(wb.sheetnames, 1):
        ws = wb[name]
        max_r, max_c = used_bounds(ws)
        state = ws.sheet_state  # visible / hidden / veryHidden
        overview_rows.append((idx, name, max_r, max_c, state))
        sheet_infos.append((idx, name, ws, max_r, max_c, state))

    lines: list[str] = [
        f"# {src.stem} (Excel extract)",
        "",
        f"- source: `{src.name}`",
        f"- extracted_with: xlsx_extract.py ({'xlsx/xlsm' if ext != '.xls' else 'xls via Excel COM'})",
        f"- sheets: {len(wb.sheetnames)}",
        f"- assets: `assets/{src.stem}/`",
        "",
        "## Sheets overview",
        "",
        "| # | name | rows | cols | state |",
        "| --- | --- | --- | --- | --- |",
    ]
    for idx, name, r, c, state in overview_rows:
        safe = name.replace("|", "\\|")
        lines.append(f"| {idx} | {safe} | {r} | {c} | {state} |")
    lines.append("")

    total_imgs = 0
    for idx, name, ws, max_r, max_c, state in sheet_infos:
        safe = name.replace("|", "\\|")
        lines.append("---")
        lines.append("")
        lines.append(f"## Sheet {idx} — {safe}")
        lines.append("")
        lines.append(f"- state: {state}")
        lines.append(f"- used range: {max_r} rows x {max_c} cols")
        if max_r == 0 or max_c == 0:
            lines.append("")
            lines.append("_(empty sheet)_")
            lines.append("")
            continue

        imgs = extract_sheet_images(ws, assets_dir, idx)
        total_imgs += len(imgs)
        if imgs:
            lines.append(f"- images: {len(imgs)}")
        lines.append("")

        lines.extend(render_sheet_table(ws, max_r, max_c))
        lines.append("")

        if imgs:
            lines.append("**Images:**")
            lines.append("")
            for n in imgs:
                lines.append(f"- ![{n}](assets/{src.stem}/{n})")
            lines.append("")

    out_md = out_dir / f"{src.stem}_extract.md"
    out_md.write_text("\n".join(lines), encoding="utf-8")
    print(f"[xlsx_extract] wrote {out_md}")
    print(f"[xlsx_extract] sheets: {len(wb.sheetnames)}, images: {total_imgs} in {assets_dir}")

    if tmp_to_clean and tmp_to_clean.exists():
        try:
            tmp_to_clean.unlink()
        except OSError:
            pass


if __name__ == "__main__":
    main()
